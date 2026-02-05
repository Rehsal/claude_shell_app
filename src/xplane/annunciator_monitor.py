"""
Annunciator Monitor - Monitors flight control annunciators and triggers alerts.

Parses sw_checklist blocks from XLSX files and continuously monitors
datarefs via ExtPlane. When all conditions in a block become true (rising edge),
triggers a tone, spoken message, and floating annunciator display.

XLSX Format (Column A):
    sw_checklist: MESSAGE
    sw_show:dataref:from_value:to_value
    sw_remark:MESSAGE

Lines starting with # or #### are ignored (comments/section headers).
"""

import logging
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

logger = logging.getLogger(__name__)


@dataclass
class AnnunciatorCondition:
    """A single dataref condition for an annunciator."""
    dataref: str
    from_value: float
    to_value: float

    def is_satisfied(self, current_value: Any) -> bool:
        """Check if the condition is satisfied by the current value."""
        if current_value is None:
            return False

        # Handle array datarefs (e.g., thrust_reverser_deploy_ratio is per-engine)
        # Always use max() for arrays:
        # - For "increasing" (EXTENDED): max >= threshold means ANY is deployed
        # - For "decreasing" (RETRACTED): max <= threshold means ALL are retracted
        try:
            if isinstance(current_value, (list, tuple)):
                if not current_value:
                    return False
                val = max(float(v) for v in current_value)
            else:
                val = float(current_value)
        except (TypeError, ValueError):
            return False

        # Determine direction and check condition
        # If to_value > from_value: condition true when value >= to_value (moving up)
        # If to_value < from_value: condition true when value <= to_value (moving down)
        # If to_value == from_value: condition true when value == to_value (exact match)
        if self.to_value > self.from_value:
            # Moving towards higher value (e.g., speedbrake extending: 0->1)
            threshold = (self.from_value + self.to_value) / 2
            return val >= threshold
        elif self.to_value < self.from_value:
            # Moving towards lower value (e.g., speedbrake retracting: 1->0)
            threshold = (self.from_value + self.to_value) / 2
            return val <= threshold
        else:
            # Exact match (e.g., rockets_armed == 0)
            return abs(val - self.to_value) < 0.01


@dataclass
class Annunciator:
    """An annunciator with conditions and message."""
    name: str
    conditions: List[AnnunciatorCondition] = field(default_factory=list)
    message: str = ""
    was_triggered: bool = False  # Track state to detect edge transitions
    last_triggered: float = 0.0
    cooldown: float = 3.0  # Minimum seconds between re-triggers


@dataclass
class AnnunciatorAlert:
    """An alert to be displayed/spoken."""
    name: str
    message: str
    timestamp: float


class AnnunciatorMonitor:
    """
    Monitors annunciator conditions and triggers alerts.

    Runs a background thread that continuously polls datarefs and
    checks annunciator conditions. When all conditions for an annunciator
    become true (rising edge), triggers an alert.
    """

    def __init__(self, extplane_client):
        self.extplane = extplane_client
        self._annunciators: Dict[str, Annunciator] = {}
        self._alerts: List[AnnunciatorAlert] = []
        self._alerts_lock = threading.Lock()
        self._max_alerts = 50  # Keep last N alerts in history
        self._loaded_path: Optional[str] = None

        # Monitoring thread
        self._running = False
        self._monitor_thread: Optional[threading.Thread] = None
        self._poll_interval = 0.3  # Poll every 300ms

        # Pending alerts queue for UI to fetch (FIFO)
        self._pending_alerts: List[AnnunciatorAlert] = []
        self._pending_lock = threading.Lock()

        # Subscribed datarefs
        self._subscribed_datarefs: set = set()

    def load_from_xlsx(self, filepath: str) -> int:
        """
        Load annunciators from an XLSX file.

        Reads column A which contains sw_checklist directives.
        Returns the number of annunciators loaded.
        """
        from openpyxl import load_workbook

        path = Path(filepath)
        if not path.exists():
            logger.warning(f"Annunciator file not found: {filepath}")
            return 0

        wb = load_workbook(filepath, data_only=True)
        try:
            ws = wb.active

            # Read all values from column A
            lines = []
            for row in ws.iter_rows(min_col=1, max_col=1):
                cell_value = row[0].value
                if cell_value:
                    lines.append(str(cell_value).strip())
                else:
                    lines.append("")

            count = self._parse_lines(lines)
            self._loaded_path = filepath
            return count
        finally:
            wb.close()

    def _parse_lines(self, lines: List[str]) -> int:
        """
        Parse annunciator definitions from lines.

        Returns the number of annunciators loaded.
        """
        current_annunciator: Optional[Annunciator] = None
        loaded_count = 0

        for line in lines:
            line = line.strip()

            # Skip empty lines, comments, and section headers
            if not line or line.startswith('#'):
                continue

            # Start of new annunciator
            if line.lower().startswith('sw_checklist:'):
                # Save previous annunciator if valid
                if current_annunciator and current_annunciator.conditions:
                    self._annunciators[current_annunciator.name] = current_annunciator
                    loaded_count += 1
                    logger.debug(f"Loaded annunciator: {current_annunciator.name}")

                # Start new annunciator
                name = line.split(':', 1)[1].strip()
                current_annunciator = Annunciator(name=name, message=name)

            # Condition line: sw_show:dataref:from:to
            elif line.lower().startswith('sw_show:') and current_annunciator:
                parts = line.split(':')
                if len(parts) >= 4:
                    dataref = parts[1].strip()
                    try:
                        from_val = float(parts[2].strip())
                        to_val = float(parts[3].strip())
                        condition = AnnunciatorCondition(
                            dataref=dataref,
                            from_value=from_val,
                            to_value=to_val
                        )
                        current_annunciator.conditions.append(condition)
                    except ValueError as e:
                        logger.warning(f"Invalid condition values in: {line} - {e}")
                elif len(parts) == 3:
                    # Format: sw_show:dataref:value (exact match)
                    dataref = parts[1].strip()
                    try:
                        val = float(parts[2].strip())
                        condition = AnnunciatorCondition(
                            dataref=dataref,
                            from_value=val,
                            to_value=val
                        )
                        current_annunciator.conditions.append(condition)
                    except ValueError as e:
                        logger.warning(f"Invalid condition value in: {line} - {e}")

            # Remark/message line
            elif line.lower().startswith('sw_remark:') and current_annunciator:
                current_annunciator.message = line.split(':', 1)[1].strip()

        # Save last annunciator
        if current_annunciator and current_annunciator.conditions:
            self._annunciators[current_annunciator.name] = current_annunciator
            loaded_count += 1
            logger.debug(f"Loaded annunciator: {current_annunciator.name}")

        logger.info(f"Loaded {loaded_count} annunciators")
        return loaded_count

    def get_annunciators(self) -> Dict[str, dict]:
        """Get all annunciators and their current state."""
        result = {}
        for name, ann in self._annunciators.items():
            result[name] = {
                'name': ann.name,
                'message': ann.message,
                'conditions': [
                    {
                        'dataref': c.dataref,
                        'from': c.from_value,
                        'to': c.to_value
                    }
                    for c in ann.conditions
                ],
                'is_active': ann.was_triggered,
                'last_triggered': ann.last_triggered
            }
        return result

    def get_pending_alert(self) -> Optional[dict]:
        """Get and remove the oldest pending alert from the queue."""
        with self._pending_lock:
            if self._pending_alerts:
                alert = self._pending_alerts.pop(0)  # FIFO - get oldest first
                return {
                    'name': alert.name,
                    'message': alert.message,
                    'timestamp': alert.timestamp
                }
        return None

    def get_alerts(self, since: float = 0) -> List[dict]:
        """Get alerts since a timestamp."""
        with self._alerts_lock:
            return [
                {
                    'name': a.name,
                    'message': a.message,
                    'timestamp': a.timestamp
                }
                for a in self._alerts
                if a.timestamp > since
            ]

    def clear_alerts(self) -> None:
        """Clear all alerts."""
        with self._alerts_lock:
            self._alerts.clear()
        with self._pending_lock:
            self._pending_alerts.clear()

    def start(self) -> bool:
        """Start the monitoring thread."""
        if self._running:
            return True

        if not self.extplane.is_connected:
            logger.warning("ExtPlane not connected, cannot start annunciator monitor")
            return False

        if not self._annunciators:
            logger.warning("No annunciators loaded, nothing to monitor")
            return False

        # Subscribe to all datarefs used by annunciators
        self._subscribe_datarefs()

        # Wait briefly for initial dataref values to arrive
        time.sleep(0.5)

        # Initialize was_triggered to current state to avoid false alerts on startup
        for ann in self._annunciators.values():
            ann.was_triggered = self._check_annunciator(ann)
            logger.debug(f"Initial state for {ann.name}: {ann.was_triggered}")

        self._running = True
        self._monitor_thread = threading.Thread(target=self._monitor_loop, daemon=True)
        self._monitor_thread.start()
        logger.info("Annunciator monitor started (initialized to current state)")
        return True

    def stop(self) -> None:
        """Stop the monitoring thread."""
        self._running = False
        if self._monitor_thread:
            self._monitor_thread.join(timeout=2.0)
            self._monitor_thread = None

        # Unsubscribe from datarefs
        self._unsubscribe_datarefs()
        logger.info("Annunciator monitor stopped")

    def is_running(self) -> bool:
        """Check if monitor is running."""
        return self._running

    def get_status(self) -> dict:
        """Get current monitor status."""
        return {
            'running': self._running,
            'loaded_path': self._loaded_path,
            'annunciator_count': len(self._annunciators),
            'subscribed_datarefs': len(self._subscribed_datarefs),
            'alert_count': len(self._alerts)
        }

    def _subscribe_datarefs(self) -> None:
        """Subscribe to all datarefs needed by annunciators."""
        datarefs = set()
        for ann in self._annunciators.values():
            for cond in ann.conditions:
                datarefs.add(cond.dataref)

        for dataref in datarefs:
            if dataref not in self._subscribed_datarefs:
                try:
                    self.extplane.subscribe(dataref, accuracy=0.01)
                    self._subscribed_datarefs.add(dataref)
                    logger.debug(f"Subscribed to: {dataref}")
                except Exception as e:
                    logger.warning(f"Failed to subscribe to {dataref}: {e}")

    def _unsubscribe_datarefs(self) -> None:
        """Unsubscribe from all datarefs."""
        for dataref in self._subscribed_datarefs:
            try:
                self.extplane.unsubscribe(dataref)
            except Exception:
                pass
        self._subscribed_datarefs.clear()

    def _get_dataref_value(self, dataref: str) -> Optional[Any]:
        """Get current value of a dataref."""
        try:
            dv = self.extplane.get_subscribed_value(dataref)
            return dv.value if dv else None
        except Exception:
            return None

    def _check_annunciator(self, ann: Annunciator) -> bool:
        """Check if all conditions for an annunciator are satisfied."""
        if not ann.conditions:
            return False

        for cond in ann.conditions:
            value = self._get_dataref_value(cond.dataref)
            if not cond.is_satisfied(value):
                return False

        return True

    def _trigger_alert(self, ann: Annunciator) -> None:
        """Trigger an alert for an annunciator."""
        now = time.time()

        # Check cooldown
        if now - ann.last_triggered < ann.cooldown:
            return

        ann.last_triggered = now

        alert = AnnunciatorAlert(
            name=ann.name,
            message=ann.message,
            timestamp=now
        )

        # Store in history
        with self._alerts_lock:
            self._alerts.append(alert)
            if len(self._alerts) > self._max_alerts:
                self._alerts = self._alerts[-self._max_alerts:]

        # Add to pending queue for UI to fetch
        with self._pending_lock:
            self._pending_alerts.append(alert)
            # Limit queue size to prevent memory issues
            if len(self._pending_alerts) > 10:
                self._pending_alerts = self._pending_alerts[-10:]

        logger.info(f"Annunciator triggered: {ann.name} - {ann.message}")

    def _monitor_loop(self) -> None:
        """Main monitoring loop."""
        logger.info("Annunciator monitor loop started")

        while self._running:
            try:
                for ann in self._annunciators.values():
                    is_satisfied = self._check_annunciator(ann)

                    # Detect rising edge (was not triggered, now is)
                    if is_satisfied and not ann.was_triggered:
                        self._trigger_alert(ann)

                    ann.was_triggered = is_satisfied

            except Exception as e:
                logger.error(f"Annunciator monitor error: {e}")

            time.sleep(self._poll_interval)

        logger.info("Annunciator monitor loop stopped")
