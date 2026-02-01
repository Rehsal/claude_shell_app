"""
Checklist Copilot - Parses XChecklist Clist.txt and automatically satisfies
checklist items by writing datarefs via ExtPlane.
"""

import csv
import io
import logging
import re
import threading
import time
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


class ItemType(Enum):
    SW_ITEM = "sw_item"
    SW_ITEMVOID = "sw_itemvoid"
    SW_REMARK = "sw_remark"
    SW_CONTINUE = "sw_continue"


class RunnerState(Enum):
    IDLE = "idle"
    RUNNING = "running"
    PAUSED = "paused"
    WAITING_CONFIRM = "waiting_confirm"
    COMPLETED = "completed"


@dataclass
class DatarefCondition:
    dataref: str
    operator: str  # =, >, <, >=, <=, !, |, +>, +<, etc.
    value: Any
    index: Optional[int] = None  # array index like [0]
    cross_ref: Optional[str] = None  # {other_dataref} reference


@dataclass
class ChecklistItem:
    item_type: ItemType
    label: str = ""
    checked_text: str = ""
    conditions: List[DatarefCondition] = field(default_factory=list)
    condition_logic: str = "&&"  # && or ||
    continue_target: str = ""  # for sw_continue
    raw_line: str = ""
    # CSV-specific fields
    group: str = ""
    order: int = 0
    speak: str = ""
    speak_on: bool = True
    command_override: str = ""
    enabled: bool = True
    wait: bool = False


@dataclass
class Checklist:
    name: str
    items: List[ChecklistItem] = field(default_factory=list)


def _parse_condition(cond_str: str) -> Optional[DatarefCondition]:
    """Parse a single dataref condition like 'sim/cockpit/electrical/battery_on[0]:1'."""
    cond_str = cond_str.strip()
    if not cond_str:
        return None

    # Check for change-detection prefix (+)
    is_trigger = False
    if cond_str.startswith('+'):
        is_trigger = True
        cond_str = cond_str[1:]

    # Split on the first operator character after the dataref path
    # Format: dataref_path[index]:operator_value or dataref_path:operator_value
    # Find the colon that separates dataref from condition (not inside {})
    colon_idx = -1
    brace_depth = 0
    for i, ch in enumerate(cond_str):
        if ch == '{':
            brace_depth += 1
        elif ch == '}':
            brace_depth -= 1
        elif ch == ':' and brace_depth == 0:
            colon_idx = i
            break

    if colon_idx == -1:
        return None

    dataref_part = cond_str[:colon_idx]
    value_part = cond_str[colon_idx + 1:]

    # Extract array index
    index = None
    idx_match = re.search(r'\[(\d+)\]', dataref_part)
    if idx_match:
        index = int(idx_match.group(1))
        dataref_part = dataref_part[:idx_match.start()]

    # Check for change-detection prefix in value_part too (e.g. dataref:+>0.5)
    if value_part.startswith('+'):
        is_trigger = True
        value_part = value_part[1:]

    # Determine operator and value
    operator = "="
    if value_part.startswith('>='):
        operator = ">=" if not is_trigger else "+>="
        value_part = value_part[2:]
    elif value_part.startswith('<='):
        operator = "<=" if not is_trigger else "+<="
        value_part = value_part[2:]
    elif value_part.startswith('>'):
        operator = ">" if not is_trigger else "+>"
        value_part = value_part[1:]
    elif value_part.startswith('<'):
        operator = "<" if not is_trigger else "+<"
        value_part = value_part[1:]
    elif value_part.startswith('!'):
        operator = "!"
        value_part = value_part[1:]
    elif value_part.startswith('|'):
        operator = "|"
        value_part = value_part[1:]
    elif '|' in value_part and not value_part.startswith('{'):
        # Range like 0.5|1.5 (pipe in the middle)
        operator = "|"
    else:
        if is_trigger:
            operator = "+="

    # Check for cross-dataref reference: {dataref} or {dataref}*multiplier
    cross_ref = None
    cross_match = re.match(r'^\{([^}]+)\}(.*)$', value_part)
    if cross_match:
        cross_ref = cross_match.group(1)
        # Store any expression suffix (e.g. "*3.28084") in value
        expr_suffix = cross_match.group(2).strip()
        value = expr_suffix if expr_suffix else None
    else:
        # Parse range for | operator
        if operator == "|" and '|' in value_part:
            # Range like 0.0|1.0 means between 0.0 and 1.0
            value = value_part  # keep as string, handle in runner
        else:
            try:
                value = float(value_part)
                if value == int(value):
                    value = int(value)
            except ValueError:
                value = value_part

    return DatarefCondition(
        dataref=dataref_part,
        operator=operator,
        value=value,
        index=index,
        cross_ref=cross_ref,
    )


def _parse_conditions(conditions_str: str) -> Tuple[List[DatarefCondition], str]:
    """Parse a conditions string, returning conditions and logic operator."""
    conditions_str = conditions_str.strip()
    if not conditions_str:
        return [], "&&"

    # Strip outer parentheses from individual conditions
    # e.g. (dataref:1)&&(dataref:2) -> dataref:1 && dataref:2

    # Determine logic: || or && (but not inside parentheses)
    if ')&&(' in conditions_str or '&&' in conditions_str:
        logic = "&&"
        parts = re.split(r'\)\s*&&\s*\(|\s*&&\s*', conditions_str)
    elif ')||(' in conditions_str or '||' in conditions_str:
        logic = "||"
        parts = re.split(r'\)\s*\|\|\s*\(|\s*\|\|\s*', conditions_str)
    else:
        logic = "&&"
        parts = [conditions_str]

    conditions = []
    for part in parts:
        # Strip surrounding parentheses
        part = part.strip().strip('()')
        cond = _parse_condition(part.strip())
        if cond:
            conditions.append(cond)

    return conditions, logic


def _parse_item_content(content: str) -> Tuple[str, str, str]:
    """
    Parse XChecklist item content: LABEL|CHECKED_TEXT:conditions

    The | separates label from checked_text. The first : after the
    checked_text (that looks like a dataref path) starts the conditions.
    Conditions contain their own : separators (dataref:value).

    Returns (label, checked_text, conditions_string).
    """
    content = content.strip()

    # Split on | to get label and rest
    pipe_idx = content.find('|')
    if pipe_idx == -1:
        # No pipe - entire content is just a label, no checked_text or conditions
        return content, "", ""

    label = content[:pipe_idx].strip()
    rest = content[pipe_idx + 1:]

    # rest = "CHECKED_TEXT:conditions" or just "CHECKED_TEXT"
    # The checked_text is human-readable (e.g. "ON", "ARMED", "SET")
    # Conditions start with a dataref path or ( character
    # Find the first : that starts a condition (preceded by what looks
    # like end of display text, followed by a dataref-like path or paren)
    # Strategy: look for :( or :[a-z] patterns which indicate conditions
    cond_start = _find_condition_start(rest)

    if cond_start == -1:
        return label, rest.strip(), ""
    else:
        checked_text = rest[:cond_start].strip()
        cond_str = rest[cond_start + 1:].strip()
        return label, checked_text, cond_str


def _find_condition_start(text: str) -> int:
    """
    Find the index of the : that separates checked_text from conditions.

    Returns -1 if no conditions found.
    """
    # Look for : followed by ( or a lowercase letter (start of dataref path)
    # or + (trigger prefix)
    # Skip : that are inside parentheses (those are dataref:value separators)
    for i, ch in enumerate(text):
        if ch == ':':
            remaining = text[i + 1:]
            # Check if what follows looks like a condition start
            if remaining and (remaining[0] == '(' or remaining[0] == '+'
                             or re.match(r'[a-z0-9]', remaining[0])):
                return i
    return -1


def parse_clist(text: str) -> List[Checklist]:
    """Parse Clist.txt content into structured checklists."""
    checklists: List[Checklist] = []
    current: Optional[Checklist] = None

    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith('//') or line.startswith('#'):
            continue

        # Checklist header: sw_checklist:NAME or sw_checklist:NAME:DISPLAY_NAME
        if line.startswith('sw_checklist:'):
            raw_name = line[len('sw_checklist:'):].strip()
            # Format can be NAME:DISPLAY_NAME - use first part as key
            name = raw_name.split(':')[0] if ':' in raw_name else raw_name
            current = Checklist(name=name)
            checklists.append(current)
            continue

        if current is None:
            continue

        # XChecklist format: sw_item:LABEL|CHECKED_TEXT:conditions
        # The | separates label from checked_text, then the first :
        # after checked_text begins the dataref conditions.
        # Conditions themselves contain : (dataref:value), so we
        # find the split point by locating the | first.
        if line.startswith('sw_item:'):
            label, checked_text, cond_str = _parse_item_content(line[len('sw_item:'):])
            conditions, logic = _parse_conditions(cond_str)
            current.items.append(ChecklistItem(
                item_type=ItemType.SW_ITEM,
                label=label,
                checked_text=checked_text,
                conditions=conditions,
                condition_logic=logic,
                raw_line=line,
            ))

        elif line.startswith('sw_itemvoid:'):
            label, checked_text, cond_str = _parse_item_content(line[len('sw_itemvoid:'):])
            current.items.append(ChecklistItem(
                item_type=ItemType.SW_ITEMVOID,
                label=label,
                checked_text=checked_text,
                raw_line=line,
            ))

        elif line.startswith('sw_remark:'):
            content = line[len('sw_remark:'):]
            current.items.append(ChecklistItem(
                item_type=ItemType.SW_REMARK,
                label=content.strip(),
                raw_line=line,
            ))

        elif line.startswith('sw_continue:'):
            content = line[len('sw_continue:'):].strip()
            # Format: TARGET or TARGET:condition
            # Find where condition starts (lowercase dataref path after :)
            cond_start = _find_condition_start(content)
            if cond_start != -1:
                target = content[:cond_start].strip()
            else:
                target = content
            current.items.append(ChecklistItem(
                item_type=ItemType.SW_CONTINUE,
                continue_target=target,
                raw_line=line,
            ))

    return checklists


# Tabular column names (shared by CSV and XLSX parsers)
_COLUMNS = [
    "checklist", "order", "group", "type", "label", "status_text",
    "speak", "speak_on", "conditions", "commands", "enabled", "wait", "notes",
]

_TYPE_MAP = {
    "item": ItemType.SW_ITEM,
    "void": ItemType.SW_ITEMVOID,
    "remark": ItemType.SW_REMARK,
    "continue": ItemType.SW_CONTINUE,
}


def _cell_str(val) -> str:
    """Normalise a cell value (from CSV or XLSX) to a stripped string."""
    if val is None:
        return ""
    return str(val).strip()


def _rows_to_checklists(rows: List[Dict[str, str]]) -> List[Checklist]:
    """Convert a list of column-keyed row dicts into Checklist objects.

    Checklists are returned in the order they first appear in the rows.
    Items within each checklist are sorted by the ``order`` column.
    """
    checklist_items: Dict[str, List[ChecklistItem]] = {}
    checklist_seen_order: List[str] = []  # preserves first-appearance order

    for row in rows:
        cl_name = _cell_str(row.get("checklist"))
        if not cl_name:
            continue

        row_type = _cell_str(row.get("type")) or "item"
        item_type = _TYPE_MAP.get(row_type.lower())
        if item_type is None:
            continue

        label = _cell_str(row.get("label"))
        status_text = _cell_str(row.get("status_text"))
        cond_str = _cell_str(row.get("conditions"))
        # Strip braces used to protect special chars in CSV
        if cond_str.startswith("{") and cond_str.endswith("}"):
            cond_str = cond_str[1:-1]
        conditions, logic = _parse_conditions(cond_str) if cond_str else ([], "&&")

        enabled_val = _cell_str(row.get("enabled")) or "1"
        enabled = enabled_val != "0"

        speak_on_val = _cell_str(row.get("speak_on")) or "1"
        speak_on = speak_on_val != "0"

        wait_val = _cell_str(row.get("wait")) or "0"
        wait = wait_val == "1"

        try:
            order = int(float(_cell_str(row.get("order")) or "0"))
        except ValueError:
            order = 0

        continue_target = label if item_type == ItemType.SW_CONTINUE else ""

        item = ChecklistItem(
            item_type=item_type,
            label=label,
            checked_text=status_text,
            conditions=conditions,
            condition_logic=logic,
            continue_target=continue_target,
            raw_line="",
            group=_cell_str(row.get("group")),
            order=order,
            speak=_cell_str(row.get("speak")),
            speak_on=speak_on,
            command_override=_cell_str(row.get("commands")),
            enabled=enabled,
            wait=wait,
        )

        if cl_name not in checklist_items:
            checklist_items[cl_name] = []
            checklist_seen_order.append(cl_name)
        checklist_items[cl_name].append(item)

    # Sort items within each checklist by order column
    for items in checklist_items.values():
        items.sort(key=lambda it: it.order)

    return [Checklist(name=n, items=checklist_items[n]) for n in checklist_seen_order]


def parse_csv_checklist(text: str) -> List[Checklist]:
    """Parse CSV checklist content into structured checklists."""
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        if not row or all(not v for v in row.values()):
            continue
        first_val = next((v for v in row.values() if v), "")
        if first_val.startswith("#"):
            continue
        rows.append(row)
    return _rows_to_checklists(rows)


def parse_xlsx_checklist(xlsx_path: str, sheet_name: str = None) -> List[Checklist]:
    """Parse an XLSX checklist workbook.

    Reads the named sheet (or first sheet) and treats row 1 as headers.
    If a named table exists on the sheet, reads from that table.
    Opens in normal (non-read-only) mode so tables are accessible and
    the file is released immediately after close.
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    try:
        # Pick the sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Try to find a named table on the sheet
        table = None
        if ws.tables:
            table = list(ws.tables.values())[0]

        rows: List[Dict[str, str]] = []

        if table:
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            header_row = [ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
            headers = [_cell_str(h).lower() for h in header_row]
            for r in range(min_row + 1, max_row + 1):
                vals = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
                if all(v is None for v in vals):
                    continue
                rows.append({headers[i]: vals[i] for i in range(len(headers))})
        else:
            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [_cell_str(h).lower() for h in header_row]
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row_cells):
                    continue
                row_dict = {}
                for i, val in enumerate(row_cells):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                rows.append(row_dict)
    finally:
        wb.close()

    return _rows_to_checklists(rows)


def _compute_write_value(cond: DatarefCondition) -> Optional[float]:
    """Determine the value to write to satisfy a condition."""
    # Skip trigger/change-detection operators
    if cond.operator.startswith('+'):
        return None

    val = cond.value
    op = cond.operator

    if op == "=":
        return float(val) if val is not None else None
    elif op == ">":
        v = float(val)
        return v + 1.0 if v != 0.0 else 1.0
    elif op == ">=":
        return float(val)
    elif op == "<":
        v = float(val)
        return v - 1.0
    elif op == "<=":
        return float(val)
    elif op == "!":
        # Not equal to val; if val is 0, write 1; otherwise write 0
        v = float(val)
        return 0.0 if v != 0.0 else 1.0
    elif op == "|":
        # Range: value is "low|high" string - write the midpoint
        if isinstance(val, str) and '|' in val:
            parts = val.split('|')
            try:
                low = float(parts[0])
                high = float(parts[1])
                return (low + high) / 2.0
            except (ValueError, IndexError):
                return None
        return float(val) if val is not None else None
    elif op == "-":
        return float(val) if val is not None else None

    return float(val) if val is not None else None


def _check_condition(cond: DatarefCondition, current_value: Any) -> bool:
    """Check if a single condition is satisfied by the current value."""
    if cond.operator.startswith('+'):
        # Trigger operators - can't check statically, assume satisfied
        return True

    if current_value is None:
        return False

    # Handle array indexing
    val = current_value
    if cond.index is not None and isinstance(val, list):
        if cond.index < len(val):
            val = val[cond.index]
        else:
            return False

    try:
        val = float(val)
    except (TypeError, ValueError):
        return False

    target = cond.value
    op = cond.operator

    if op == "=":
        return abs(val - float(target)) < 0.05
    elif op == ">":
        return val > float(target)
    elif op == ">=":
        return val >= float(target)
    elif op == "<":
        return val < float(target)
    elif op == "<=":
        return val <= float(target)
    elif op == "!":
        return abs(val - float(target)) > 0.05
    elif op == "|":
        if isinstance(target, str) and '|' in target:
            parts = target.split('|')
            try:
                low = float(parts[0])
                high = float(parts[1])
                return low <= val <= high
            except (ValueError, IndexError):
                return False
        return abs(val - float(target)) < 0.05

    return False


class ChecklistRunner:
    """
    Runs parsed checklists using the commands system and ExtPlane.

    Uses Clist.txt as a template for sequencing. For each sw_item:
    1. Execute the matching command (e.g. "BATTERY ON") via ScriptExecutor
    2. Poll the item's dataref conditions to detect completion
    3. Advance automatically when conditions are met

    Runs in a background thread so the API stays responsive.
    Uses threading.Event for pause/resume/confirm/skip signaling.
    """

    def __init__(self, client, commands_loader=None, poll_timeout: float = 8.0):
        self.client = client
        self.commands_loader = commands_loader
        self.checklists: List[Checklist] = []
        self._checklist_map: Dict[str, Checklist] = {}
        self.current_checklist: Optional[Checklist] = None
        self.current_index: int = 0
        self.state: RunnerState = RunnerState.IDLE
        self._log: List[str] = []
        self._poll_timeout: float = poll_timeout

        # History tracking — every processed item gets appended here
        self._history: List[Dict] = []
        self._current_section: str = ""
        self._progress_total: int = 0  # items processed across all sub-checklists

        # Threading
        self._thread: Optional[threading.Thread] = None
        self._stop_event = threading.Event()
        self.auto_continue: bool = True
        self.speech_enabled: bool = True
        self._speech_event = threading.Event()
        self._speech_event.set()  # start unblocked
        self._pause_event = threading.Event()  # set = running, clear = paused
        self._confirm_event = threading.Event()
        self._skip_event = threading.Event()
        self._pause_event.set()  # start unpaused
        self._lock = threading.Lock()

    def load(self, clist_path: str) -> List[str]:
        self.stop()
        path = Path(clist_path)
        if not path.exists():
            raise FileNotFoundError(f"Checklist file not found: {clist_path}")
        ext = path.suffix.lower()
        if ext == '.xlsx':
            self.checklists = parse_xlsx_checklist(str(path))
        elif ext == '.csv':
            text = path.read_text(encoding='utf-8', errors='replace')
            self.checklists = parse_csv_checklist(text)
        else:
            text = path.read_text(encoding='utf-8', errors='replace')
            self.checklists = parse_clist(text)
        self._checklist_map = {cl.name: cl for cl in self.checklists}
        total_items = sum(len(cl.items) for cl in self.checklists)
        from datetime import datetime
        mtime = datetime.fromtimestamp(path.stat().st_mtime).strftime("%I:%M:%S %p")
        self._load_summary = f"{len(self.checklists)} checklists, {total_items} items — file saved {mtime}"
        print(f"[checklist] Loaded {self._load_summary}")
        self.state = RunnerState.IDLE
        self.current_checklist = None
        self.current_index = 0
        self._log = []
        self._history = []
        self._current_section = ""
        self._progress_total = 0
        return [cl.name for cl in self.checklists]

    def list_checklists(self) -> List[Dict]:
        return [{"name": cl.name, "item_count": len(cl.items)}
                for cl in self.checklists]

    def start(self, checklist_name: str, auto_run: bool = False) -> bool:
        """Start a checklist. If auto_run=True, begins executing in background."""
        cl = self._checklist_map.get(checklist_name)
        if not cl:
            return False
        self.stop()
        self.current_checklist = cl
        self.current_index = 0
        self.state = RunnerState.RUNNING
        self._current_section = checklist_name
        self._history = []
        self._progress_total = 0
        self._log.append(f"Started checklist: {checklist_name}")
        if auto_run:
            self._start_background()
        return True

    def stop(self):
        """Stop background execution."""
        self._stop_event.set()
        self._pause_event.set()  # unblock if paused
        self._confirm_event.set()  # unblock if waiting
        if self._thread and self._thread.is_alive():
            self._thread.join(timeout=5)
        self._thread = None
        self._stop_event.clear()
        self._confirm_event.clear()
        self._skip_event.clear()
        self._pause_event.set()

    def restart(self):
        """Reset to START checklist and re-run from the beginning."""
        self.stop()
        self._history = []
        self._progress_total = 0
        self._current_section = ""
        self._log = []
        # Find the START checklist (first one whose name contains 'START', or just the first)
        start_cl = None
        for cl in self.checklists:
            if 'START' in cl.name.upper():
                start_cl = cl
                break
        if not start_cl and self.checklists:
            start_cl = self.checklists[0]
        if start_cl:
            self.current_checklist = start_cl
            self.current_index = 0
            self._current_section = start_cl.name
            self.state = RunnerState.RUNNING
            self._log.append(f"Restarted from: {start_cl.name}")
            self._start_background()

    def pause(self):
        if self.state == RunnerState.RUNNING:
            self._pause_event.clear()
            self.state = RunnerState.PAUSED
            self._log.append("Paused")

    def resume(self):
        if self.state in (RunnerState.PAUSED, RunnerState.WAITING_CONFIRM):
            self.state = RunnerState.RUNNING
            self._pause_event.set()
            self._confirm_event.set()
            self._log.append("Resumed")

    def confirm(self):
        """Confirm a manual item."""
        if self.state == RunnerState.WAITING_CONFIRM:
            self._log.append(f"Confirmed: {self._current_item_label()}")
            self._confirm_event.set()

    def skip_item(self):
        """Skip the current item."""
        self._log.append(f"Skipped: {self._current_item_label()}")
        self._skip_event.set()
        # Also unblock confirm wait
        self._confirm_event.set()

    def run_single(self) -> Dict:
        """Run just the current item (blocking). For step-by-step mode."""
        return self._process_current_item()

    def run_all(self):
        """Start auto-running all items in a background thread."""
        if self.state != RunnerState.RUNNING:
            if self.current_checklist:
                self.state = RunnerState.RUNNING
        self._start_background()

    def _start_background(self):
        """Launch the background runner thread."""
        self._stop_event.clear()
        self._pause_event.set()
        self._thread = threading.Thread(target=self._run_loop, daemon=True,
                                        name="checklist-runner")
        self._thread.start()
        self._log.append("Auto-run started")

    def _ensure_connected(self, timeout: float = 30.0) -> bool:
        """Wait for ExtPlane client to connect, retrying if needed."""
        start = time.time()
        while time.time() - start < timeout:
            if self._stop_event.is_set():
                return False
            if self.client.is_connected:
                return True
            if not self.client.is_connected:
                self.client.connect()
            time.sleep(1.0)
        return self.client.is_connected

    def _run_loop(self):
        """Background thread: process items until done, paused, or stopped."""
        if not self._ensure_connected():
            self._log.append("ExtPlane not connected — aborting checklist")
            self.state = RunnerState.IDLE
            return

        while not self._stop_event.is_set():
            # Wait if paused
            self._pause_event.wait(timeout=1.0)
            if self._stop_event.is_set():
                break

            if self.state not in (RunnerState.RUNNING,):
                if self.state == RunnerState.COMPLETED:
                    break
                time.sleep(0.5)
                continue

            try:
                result = self._process_current_item()
            except Exception as e:
                logger.error(f"Runner error: {e}")
                self._log.append(f"Error: {e}")
                self._advance()
                continue

            action = result.get("action", "")
            if action == "completed":
                break
            if action == "none":
                time.sleep(0.5)
            else:
                pass  # no artificial delay

        self._log.append("Auto-run finished")

    def _current_item_label(self) -> str:
        if self.current_checklist and self.current_index < len(self.current_checklist.items):
            item = self.current_checklist.items[self.current_index]
            return item.label or item.continue_target or "(unknown)"
        return ""

    def _append_history(self, item: ChecklistItem, status: str):
        """Append a processed item to the history list."""
        entry: Dict[str, Any] = {
            "label": item.label,
            "checked_text": item.checked_text,
            "type": item.item_type.value,
            "status": status,
            "section": self._current_section,
            "continue_target": item.continue_target if item.item_type == ItemType.SW_CONTINUE else "",
        }
        if self.speech_enabled and item.speak_on and item.speak:
            entry["speak"] = item.speak
        self._history.append(entry)
        self._progress_total += 1
        # Block runner until client-side TTS signals completion
        if entry.get("speak"):
            self._speech_event.clear()
            # Timeout based on word count as fallback (generous to avoid cutoff)
            word_count = len(entry["speak"].split())
            timeout = max(4.0, word_count * 0.8)
            self._speech_event.wait(timeout=timeout)

    def speech_done(self):
        """Called by the API when client-side TTS finishes an utterance."""
        self._speech_event.set()

    def get_status(self) -> Dict:
        result: Dict[str, Any] = {
            "state": self.state.value,
            "current_checklist_name": self.current_checklist.name if self.current_checklist else None,
            # Keep legacy key for compat
            "checklist": self.current_checklist.name if self.current_checklist else None,
            "item_index": self.current_index,
            "total_items": len(self.current_checklist.items) if self.current_checklist else 0,
            "current_item": None,
            "progress": 0.0,
            "running_in_background": self._thread is not None and self._thread.is_alive(),
            "history": list(self._history),
            "progress_total": self._progress_total,
            "log": self._log[-100:],
            "auto_continue": self.auto_continue,
            "speech_enabled": self.speech_enabled,
        }
        if self.current_checklist and self.current_index < len(self.current_checklist.items):
            item = self.current_checklist.items[self.current_index]
            result["current_item"] = {
                "type": item.item_type.value,
                "label": item.label,
                "checked_text": item.checked_text,
                "has_datarefs": len(item.conditions) > 0,
                "continue_target": item.continue_target,
            }
            total = len(self.current_checklist.items)
            if total > 0:
                result["progress"] = round(self.current_index / total * 100, 1)
        return result

    # ----------------------------------------------------------------
    # Item processing
    # ----------------------------------------------------------------

    def _advance(self):
        self.current_index += 1
        if self.current_checklist and self.current_index >= len(self.current_checklist.items):
            self.state = RunnerState.COMPLETED

    def _process_current_item(self) -> Dict:
        """Process the current item. May block while polling conditions."""
        if self.state not in (RunnerState.RUNNING,):
            return {"action": "none", "reason": f"Runner is {self.state.value}"}

        # Ensure ExtPlane is connected before processing any item
        if not self.client.is_connected:
            if not self._ensure_connected(timeout=10.0):
                return {"action": "error", "reason": "ExtPlane not connected"}

        if not self.current_checklist:
            return {"action": "none", "reason": "No checklist loaded"}

        if self.current_index >= len(self.current_checklist.items):
            self.state = RunnerState.COMPLETED
            return {"action": "completed"}

        item = self.current_checklist.items[self.current_index]

        # Skip disabled items (CSV enabled=0)
        if not item.enabled:
            self._log.append(f"[disabled] {item.label}")
            self._advance()
            return {"action": "skipped", "reason": "disabled", "label": item.label}

        # sw_itemvoid / sw_remark - display only, advance
        if item.item_type in (ItemType.SW_ITEMVOID, ItemType.SW_REMARK):
            self._log.append(f"[{item.item_type.value}] {item.label}")
            self._append_history(item, "void" if item.item_type == ItemType.SW_ITEMVOID else "remark")
            self._advance()
            return {"action": "skipped", "type": item.item_type.value,
                    "label": item.label}

        # sw_continue - chain to next checklist
        if item.item_type == ItemType.SW_CONTINUE:
            target = item.continue_target
            if not self.auto_continue:
                self._log.append(f"[auto-continue off] skipping >> {target}")
                self._append_history(item, "skipped")
                self._advance()
                return {"action": "skipped", "type": "sw_continue", "label": item.label}
            self._log.append(f">> {target}")
            self._append_history(item, "continue")
            cl = self._checklist_map.get(target)
            if cl:
                self.current_checklist = cl
                self.current_index = 0
                self._current_section = cl.name
                return {"action": "continue", "target": target}
            else:
                self._advance()
                return {"action": "error", "reason": f"Checklist '{target}' not found"}

        # sw_item - the main work
        if item.item_type == ItemType.SW_ITEM:
            # If there's a command override but no conditions, execute and advance
            if not item.conditions and item.command_override:
                self._log.append(f"Command-only: {item.label}")
                cmd_result = self._try_execute_command(item)
                status = "executed" if cmd_result else "no_match"
                self._append_history(item, status)
                self._advance()
                return {"action": status, "label": item.label}
            if not item.conditions:
                # No conditions — try to execute command by label, then advance
                self._log.append(f"No-condition: {item.label}")
                cmd_result = self._try_execute_command(item)
                status = "executed" if cmd_result else "no_match"
                self._append_history(item, status)
                self._advance()
                return {"action": status, "label": item.label}
            return self._handle_dataref_item(item)

        return {"action": "none", "reason": "Unknown item type"}

    def _handle_manual_item(self, item: ChecklistItem) -> Dict:
        """Manual item with no datarefs - wait for pilot confirmation."""
        self.state = RunnerState.WAITING_CONFIRM
        self._log.append(f"Manual: {item.label}|{item.checked_text}")
        self._confirm_event.clear()
        self._skip_event.clear()

        # Block until confirmed, skipped, or stopped
        while not self._stop_event.is_set():
            if self._confirm_event.wait(timeout=1.0):
                break
            if self._skip_event.is_set():
                break

        self._confirm_event.clear()
        skipped = self._skip_event.is_set()
        self._skip_event.clear()

        if self._stop_event.is_set():
            return {"action": "none", "reason": "Stopped"}

        self._append_history(item, "skipped" if skipped else "confirmed")
        self._advance()
        self.state = RunnerState.RUNNING
        return {"action": "confirmed" if not skipped else "skipped",
                "label": item.label, "checked_text": item.checked_text}

    def _handle_dataref_item(self, item: ChecklistItem) -> Dict:
        """Execute command for item, then poll conditions."""
        try:
            return self._handle_dataref_item_inner(item)
        except Exception as e:
            logger.error(f"Error on {item.label}: {e}")
            self._log.append(f"Error: {item.label} - {e}")
            self._append_history(item, "error")
            self._advance()
            return {"action": "error", "label": item.label, "reason": str(e)}

    def _handle_dataref_item_inner(self, item: ChecklistItem) -> Dict:
        # Check ExtPlane connection
        if not self.client.is_connected:
            self._log.append(f"No connection - skipping: {item.label}")
            self._append_history(item, "skipped")
            self._advance()
            return {"action": "skipped", "label": item.label,
                    "reason": "ExtPlane not connected"}

        # Skip trigger-only items
        non_trigger = [c for c in item.conditions if not c.operator.startswith('+')]
        if not non_trigger:
            self._log.append(f"Trigger-only: {item.label}|{item.checked_text}")
            self._append_history(item, "skipped")
            self._advance()
            return {"action": "skipped", "type": "trigger_only",
                    "label": item.label, "checked_text": item.checked_text}

        # Determine if all non-trigger conditions are read-only (sim state)
        non_trigger = [c for c in item.conditions if not c.operator.startswith('+')]
        all_read_only = all(self._is_read_only(c.dataref) for c in non_trigger) if non_trigger else False

        # Already satisfied?
        if self._check_all_conditions(item.conditions, item.condition_logic):
            self._log.append(f"Already OK: {item.label}|{item.checked_text}")
            # For read-only conditions already met, fire command now
            if all_read_only and item.command_override:
                self._try_execute_command(item)
            self._append_history(item, "satisfied")
            self._advance()
            return {"action": "already_met", "label": item.label,
                    "checked_text": item.checked_text}

        actions = []
        errors = []

        if all_read_only:
            # Read-only conditions (N2, indicators): wait FIRST, command AFTER
            self._log.append(f"Waiting: {item.label}|{item.checked_text}")
            for cond in non_trigger:
                raw = self._read_dataref(cond.dataref, cond.index)
                met = _check_condition(cond, raw)
                idx_str = f"[{cond.index}]" if cond.index is not None else ""
                self._log.append(f"  {cond.dataref}{idx_str} {cond.operator} {cond.value}: raw={raw} met={met}")

            default_timeout = 120.0
            item_timeout = self._LONG_POLL_LABELS.get(item.label.strip(), default_timeout)
            elapsed = self._poll_until_satisfied(item, timeout_override=item_timeout)
            verified = self._check_all_conditions(item.conditions, item.condition_logic)

            if self._skip_event.is_set():
                self._skip_event.clear()
                self._log.append(f"Skipped: {item.label}")
                self._append_history(item, "skipped")
                self._advance()
                self.state = RunnerState.RUNNING
                return {"action": "skipped", "label": item.label}

            if verified:
                # Conditions met — NOW fire the command
                cmd_result = self._try_execute_command(item)
                if cmd_result:
                    actions = cmd_result.get("commands_sent", []) + cmd_result.get("datarefs_set", [])
                wait_str = f" ({elapsed:.0f}s)" if elapsed > 1 else ""
                self._log.append(f"Done{wait_str}: {item.label}|{item.checked_text}")
                self._append_history(item, "satisfied")
                self._advance()
                return {"action": "satisfied", "label": item.label,
                        "checked_text": item.checked_text, "actions": actions,
                        "wait_time": round(elapsed, 1)}
            else:
                if item.wait:
                    # wait=1: block until condition met or user skips
                    self._log.append(f"Waiting (wait=1): {item.label}|{item.checked_text}")
                    return self._handle_manual_item(item)
                self._log.append(f"Timeout — skipping: {item.label}|{item.checked_text}")
                self._append_history(item, "timeout")
                self._advance()
                return {"action": "timeout", "label": item.label,
                        "checked_text": item.checked_text}

        # Writable conditions: execute command FIRST to actuate, then poll to verify
        cmd_result = self._try_execute_command(item)
        if cmd_result:
            actions = cmd_result.get("commands_sent", []) + cmd_result.get("datarefs_set", [])
            errors = cmd_result.get("errors", [])

        # Write action datarefs (switch positions) — never status datarefs
        write_actions, write_errors = self._write_conditions(item.conditions)
        actions.extend(write_actions)
        errors.extend(write_errors)

        # Apply Zibo-specific extra writes for items where the command is incomplete
        self._zibo_post_command(item)

        # Toggle covers that are in the wrong position
        self._try_toggle_unsatisfied(item.conditions)

        # Brief pause to let command take effect
        time.sleep(0.3)

        # Log condition state before polling
        self._log.append(f"Executing: {item.label}|{item.checked_text}")
        for cond in item.conditions:
            if not cond.operator.startswith('+'):
                raw = self._read_dataref(cond.dataref, cond.index)
                met = _check_condition(cond, raw)
                idx_str = f"[{cond.index}]" if cond.index is not None else ""
                self._log.append(f"  {cond.dataref}{idx_str} {cond.operator} {cond.value}: raw={raw} met={met}")

        # Early check: if checklist conditions already met, skip polling
        if self._check_all_conditions(item.conditions, item.condition_logic):
            self._log.append(f"Done: {item.label}|{item.checked_text}")
            self._append_history(item, "satisfied")
            self._advance()
            return {"action": "satisfied", "label": item.label,
                    "checked_text": item.checked_text, "actions": actions}

        item_timeout = self._LONG_POLL_LABELS.get(item.label.strip(), self._poll_timeout)
        elapsed = self._poll_until_satisfied(item, timeout_override=item_timeout)

        verified = self._check_all_conditions(item.conditions, item.condition_logic)

        if self._skip_event.is_set():
            self._skip_event.clear()
            self._log.append(f"Skipped: {item.label}")
            self._append_history(item, "skipped")
            self._advance()
            self.state = RunnerState.RUNNING
            return {"action": "skipped", "label": item.label}

        if verified:
            wait_str = f" ({elapsed:.0f}s)" if elapsed > 1 else ""
            self._log.append(f"Done{wait_str}: {item.label}|{item.checked_text}")
            self._append_history(item, "satisfied")
            self._advance()
            return {"action": "satisfied", "label": item.label,
                    "checked_text": item.checked_text, "actions": actions,
                    "wait_time": round(elapsed, 1)}
        else:
            if item.wait:
                self._log.append(f"Waiting (wait=1): {item.label}|{item.checked_text}")
                return self._handle_manual_item(item)
            self._log.append(f"Timeout — skipping: {item.label}|{item.checked_text}")
            self._append_history(item, "timeout")
            self._advance()
            return {"action": "timeout", "label": item.label,
                    "checked_text": item.checked_text, "actions": actions}

    def _verify_command_datarefs(self, cmd_result: Dict) -> bool:
        """Check if the command's own script datarefs were set to their intended values."""
        datarefs_set = cmd_result.get("datarefs_set", [])
        if not datarefs_set:
            self._log.append(f"  cmd fallback: no datarefs_set in command result")
            return False
        for entry in datarefs_set:
            dr = entry.get("dataref", "")
            target = entry.get("value")
            if not dr or target is None:
                continue
            # Strip array index for reading
            raw_dr = dr
            idx = None
            idx_match = re.search(r'\[(\d+)\]', dr)
            if idx_match:
                idx = int(idx_match.group(1))
                raw_dr = dr[:idx_match.start()]
            current = self._read_dataref(raw_dr, idx)
            self._log.append(f"  cmd fallback: {dr} target={target} current={current}")
            if current is None:
                return False
            try:
                if abs(float(current) - float(target)) > 0.05:
                    return False
            except (TypeError, ValueError):
                return False
        return True

    def _poll_until_satisfied(self, item: ChecklistItem, timeout_override: float = None) -> float:
        """
        Poll conditions until met, skipped, paused, or timed out.
        Returns elapsed seconds.
        """
        max_wait = timeout_override if timeout_override is not None else self._poll_timeout
        poll_interval = 0.5
        elapsed = 0.0

        time.sleep(0.5)
        if self._check_all_conditions(item.conditions, item.condition_logic):
            return 0.0

        while elapsed < max_wait:
            if self._stop_event.is_set() or self._skip_event.is_set():
                break
            # Respect pause
            self._pause_event.wait(timeout=1.0)
            if self.state == RunnerState.PAUSED:
                continue

            time.sleep(poll_interval)
            elapsed += poll_interval

            if self._check_all_conditions(item.conditions, item.condition_logic):
                return elapsed

            # Slow down after initial burst
            if elapsed >= 3.0:
                poll_interval = 2.0

        return elapsed

    # ----------------------------------------------------------------
    # Condition checking & command execution
    # ----------------------------------------------------------------

    def _check_all_conditions(self, conditions: List[DatarefCondition],
                              logic: str = "&&") -> bool:
        non_trigger = [c for c in conditions if not c.operator.startswith('+')]
        if not non_trigger:
            return True
        for cond in non_trigger:
            resolved = self._resolve_cross_ref(cond)
            current = self._read_dataref(cond.dataref, cond.index)
            met = _check_condition(resolved, current)
            logger.debug(f"Condition: {cond.dataref}[{cond.index}] {cond.operator} {cond.value} | raw={current} | met={met}")
            if logic == "||" and met:
                return True
            if logic == "&&" and not met:
                return False
        return logic == "&&"

    def _resolve_cross_ref(self, cond: DatarefCondition) -> DatarefCondition:
        if not cond.cross_ref:
            return cond
        ref_val = self._read_dataref(cond.cross_ref)
        if ref_val is None:
            return cond
        resolved = float(ref_val)
        if cond.value and isinstance(cond.value, str):
            try:
                expr = cond.value
                if expr.startswith('*'):
                    resolved *= float(expr[1:])
                elif expr.startswith('/'):
                    resolved /= float(expr[1:])
                elif expr.startswith('+'):
                    resolved += float(expr[1:])
                elif expr.startswith('-'):
                    resolved -= float(expr[1:])
            except (ValueError, ZeroDivisionError):
                pass
        return DatarefCondition(
            dataref=cond.dataref, operator=cond.operator,
            value=resolved, index=cond.index,
        )

    # Known mappings from checklist labels to command text.
    # These handle cases where Clist.txt labels don't match commands.xml token phrases.
    # Each entry maps LABEL -> {CHECKED_TEXT_KEYWORD: "command phrase"}
    _LABEL_TO_COMMAND = {
        # Electrical
        "BATTERY": {"ON": "battery on", "OFF": "battery off"},
        "APU GENERATORS": {"ON": "apu generator on", "OFF": "apu generator off"},
        # Fuel
        "FUEL PUMPS LEFT AND RIGHT": {"ON": None},  # handled by _ZIBO_POST_COMMAND
        "FUEL PUMPS CENTER": {"ON": None},  # handled by _ZIBO_POST_COMMAND
        # Ice protection
        "WINDOW HEATERS": {"ON": None},  # handled by _ZIBO_POST_COMMAND
        "WINDOW HEATER": {"ON": None},  # handled by _ZIBO_POST_COMMAND
        # Lights
        "ANTI COLLISION LIGHTS": {"ON": "beacon on", "OFF": "beacon off"},
        # Signs — no commands exist; rely on direct dataref writes from conditions
        "SEATBELTS SIGN": {"ON": "seat belts on", "AUTO": "seat belts auto"},
        # Flight controls
        "YAW DAMPER": {"ON": "yaw damper on", "OFF": "yaw damper off"},
        "AUTO THROTTLE": {"ARMED": "autothrottle arm"},
        "AUTOPILOT DISENGAGE bar": {"UP": None},  # None = no command, check only
        # IRS alignment — no command, just wait for sim to complete alignment
        "IRS ALIGNMENT IS COMPLETE": {"CHECK": None},
        # Engine start
        "ENGINE 1 START SWITCH": {"GROUND": None},  # handled by _ZIBO_POST_COMMAND
        "ENGINE 2 START SWITCH": {"GROUND": None},  # handled by _ZIBO_POST_COMMAND
        "ENGINE 2 START SWITCH": {"GROUND": None},  # Clist uses lowercase 'switch'
        "ENGINE 1 START LEVER": {"IDLE": None},  # handled by _ZIBO_POST_COMMAND
        "ENGINE 2 START LEVER": {"IDLE": None},  # handled by _ZIBO_POST_COMMAND
        "ENGINE 1": {"STABLE IDLE": None},  # prevent ENGINE_ONE SHUTDOWN match
        "ENGINE 2": {"STABLE IDLE": None},  # prevent ENGINE_TWO SHUTDOWN match
        "N2 REACHING 25%": {"CHECK": None},  # read-only, just poll
        # Emergency exit
        "EMERGENCY EXIT LIGHTS": {"ARMED": None},  # handled by _ZIBO_POST_COMMAND
        # Hydraulic pump — no command in commands.xml, handled by _ZIBO_POST_COMMAND
        "SYSTEM B ELECTRIC HYDRAULIC PUMP": {"ON": None},
        "SYSTEM A ELECTRIC HYDRAULIC PUMP": {"ON": None},
        # Autobrake
        "AUTO BRAKE SELECTOR": {"RTO": "auto brakes rejected takeoff"},
        # Engine generators — label doesn't match "GENERATOR ON" token
        "ENGINE GENERATORS": {"ON": "generator on", "OFF": "generator off"},
    }

    # Items that need longer poll time (APU startup, engine spool, etc.)
    _LONG_POLL_LABELS = {
        "APU START": 120.0,      # APU takes ~60s to start
        "ENGINE 1": 60.0,        # Engine spool to stable idle
        "ENGINE 2": 60.0,
        "N2 REACHING 25%": 30.0, # Engine N2 spool
        "IRS alignment is complete": 300.0,  # IRS alignment can take minutes
        "close all Doors": 30.0,
    }

    def _guess_command_text(self, item: ChecklistItem) -> List[str]:
        label = item.label.strip()
        checked = item.checked_text.strip()
        candidates = []

        # Check known label-to-command mappings first
        label_upper = label.upper()
        if label_upper in self._LABEL_TO_COMMAND:
            mapping = self._LABEL_TO_COMMAND[label_upper]
            checked_first = checked.split(',')[0].split('or')[0].split('and')[0].strip().upper()
            if checked_first in mapping:
                if mapping[checked_first] is None:
                    # Explicitly no command — return empty, don't fall through
                    return []
                candidates.append(mapping[checked_first])
            # Also try full checked text
            checked_upper = checked.strip().upper()
            if checked_upper in mapping:
                if mapping[checked_upper] is None:
                    return []
                if mapping[checked_upper] not in candidates:
                    candidates.append(mapping[checked_upper])
        # Most specific first: label + first word of checked_text
        if label and checked:
            first_word = checked.split(',')[0].split('or')[0].split('and')[0].strip()
            if first_word:
                candidates.append(f"{label} {first_word}")
        # Label + full checked_text
        if label and checked:
            candidates.append(f"{label} {checked}")
        # First word of label + checked_text
        # (e.g. "IRS mode selectors" + "NAV" -> "IRS NAV")
        if label and checked:
            label_first = label.split()[0]
            checked_first = checked.split(',')[0].split('or')[0].split('and')[0].strip()
            short = f"{label_first} {checked_first}"
            if short not in candidates:
                candidates.append(short)
        # Label without last word + checked_text
        # (e.g. "EMERGENCY EXIT LIGHTS" + "ARMED" -> "EMERGENCY EXIT ARMED")
        label_words = label.split()
        if label and checked and len(label_words) > 1:
            checked_first = checked.split(',')[0].split('or')[0].split('and')[0].strip()
            trimmed = ' '.join(label_words[:-1]) + ' ' + checked_first
            if trimmed not in candidates:
                candidates.append(trimmed)
        # Dataref-hint: extract keywords from condition datarefs to find better commands
        # e.g. condition "mixture_ratio1" -> try "MIXTURE ONE <checked_text>"
        if item.conditions and checked:
            checked_first = checked.split(',')[0].split('or')[0].split('and')[0].strip()
            for cond in item.conditions:
                if cond.operator.startswith('+'):
                    continue
                # Extract the last path segment and look for keywords
                parts = cond.dataref.rsplit('/', 1)[-1]
                # Map engine numbers: 1/2 -> ONE/TWO
                engine_map = {'1': 'ONE', '2': 'TWO', '3': 'THREE', '4': 'FOUR'}
                for digit, word in engine_map.items():
                    if digit in parts:
                        # e.g. "mixture_ratio1" -> "MIXTURE ONE IDLE"
                        base = re.sub(r'[_\d]+', ' ', parts).strip().upper()
                        base_words = base.split()
                        if base_words:
                            dr_candidate = f"{base_words[0]} {word} {checked_first}"
                            if dr_candidate not in candidates:
                                candidates.append(dr_candidate)
        # Bare label last (least specific - may match wrong variant)
        if label:
            candidates.append(label)
        return candidates

    def _execute_single_command(self, part: str) -> Optional[Dict]:
        """Execute a single command string (set:, cmd:, or phrase lookup)."""
        from .script_executor import ScriptExecutor

        part = part.strip()
        if not part:
            return None

        if part.startswith("fms:"):
            # FMS programmer integration: fms:program_all, fms:route, etc.
            from .fms_programmer import FMSProgrammer
            fms_page = part[4:].strip()
            # Get or create FMS programmer (reuses the global instance via app.py)
            # For checklist context, we create a local one if needed
            if not hasattr(self, '_fms_programmer') or self._fms_programmer is None:
                self._fms_programmer = FMSProgrammer(self.client)
            fms = self._fms_programmer
            self._log.append(f"FMS: {fms_page}")
            success = fms.run_sync(fms_page)
            return {
                "command": part,
                "commands_sent": [f"fms:{fms_page}"],
                "datarefs_set": [],
                "errors": [] if success else [f"FMS {fms_page} failed"],
            }

        if part.startswith("set:"):
            # Direct dataref write: set:dataref=value
            try:
                dr, val_str = part[4:].split("=", 1)
                dr = dr.strip()
                val_raw = val_str.strip()
                if '.' in val_raw:
                    val = float(val_raw)
                else:
                    try:
                        val = int(val_raw)
                    except ValueError:
                        val = float(val_raw)
                self._log.append(f"Set dataref: {dr} = {val}")
                self.client.set_dataref(dr, val)
                return {
                    "command": part,
                    "commands_sent": [],
                    "datarefs_set": [{"dataref": dr, "value": val}],
                    "errors": [],
                }
            except Exception as e:
                self._log.append(f"Set dataref failed: {part} - {e}")
                return {
                    "command": part,
                    "commands_sent": [],
                    "datarefs_set": [],
                    "errors": [str(e)],
                }

        if part.startswith("toggle:"):
            # toggle:command_path=dataref_path:desired_value
            # Reads the dataref; if not at desired value, fires the toggle command.
            try:
                rest = part[7:].strip()
                cmd_path, dr_spec = rest.split("=", 1)
                cmd_path = cmd_path.strip()
                dr_path, val_str = dr_spec.rsplit(":", 1)
                dr_path = dr_path.strip()
                val_str = val_str.strip()
                desired = float(val_str) if '.' in val_str else int(val_str)

                current = self.client.get_dataref(dr_path)
                if current is not None and float(current) == float(desired):
                    self._log.append(f"Toggle skip: {dr_path} already {desired}")
                    return {
                        "command": cmd_path,
                        "commands_sent": [],
                        "datarefs_set": [],
                        "errors": [],
                    }
                self._log.append(f"Toggle: {cmd_path} ({dr_path}: {current} -> {desired})")
                self.client.send_command(cmd_path)
                return {
                    "command": cmd_path,
                    "commands_sent": [cmd_path],
                    "datarefs_set": [],
                    "errors": [],
                }
            except Exception as e:
                self._log.append(f"Toggle failed: {part} - {e}")
                return {
                    "command": part,
                    "commands_sent": [],
                    "datarefs_set": [],
                    "errors": [str(e)],
                }

        if part.startswith("cmdhold:"):
            # Hold a command briefly using begin/end (for rotary switches, etc.)
            import time
            cmd_key = part[8:].strip()
            hold_time = 0.2  # default hold duration in seconds
            # Allow optional duration: cmdhold:0.5:sim/some/command
            if ':' in cmd_key:
                maybe_time, _, rest = cmd_key.partition(':')
                try:
                    hold_time = float(maybe_time)
                    cmd_key = rest.strip()
                except ValueError:
                    pass  # not a number, treat whole thing as command path
            self._log.append(f"Command (hold {hold_time}s): {cmd_key}")
            try:
                self.client.send_command_begin(cmd_key)
                time.sleep(hold_time)
                self.client.send_command_end(cmd_key)
                return {
                    "command": cmd_key,
                    "commands_sent": [cmd_key],
                    "datarefs_set": [],
                    "errors": [],
                }
            except Exception as e:
                return {
                    "command": cmd_key,
                    "commands_sent": [],
                    "datarefs_set": [],
                    "errors": [str(e)],
                }

        if part.startswith("cmd:"):
            cmd_key = part[4:].strip()
            # If it contains '/' it's a raw X-Plane command path
            if '/' in cmd_key:
                self._log.append(f"Command (raw): {cmd_key}")
                try:
                    self.client.send_command(cmd_key)
                    return {
                        "command": cmd_key,
                        "commands_sent": [cmd_key],
                        "datarefs_set": [],
                        "errors": [],
                    }
                except Exception as e:
                    return {
                        "command": cmd_key,
                        "commands_sent": [],
                        "datarefs_set": [],
                        "errors": [str(e)],
                    }
            # Otherwise lookup by command key in commands.xml
            if self.commands_loader:
                cmd = self.commands_loader.get_command(cmd_key.upper())
                if cmd and cmd.script:
                    self._log.append(f"Command (override key): {cmd_key}")
                    executor = ScriptExecutor(self.client)
                    result = executor.execute(cmd.script, {})
                    return {
                        "command": cmd_key,
                        "commands_sent": result.get("commands_sent", []),
                        "datarefs_set": result.get("datarefs_set", []),
                        "errors": result.get("errors", []),
                    }
            return None

        # Exact phrase — use as the sole candidate
        if self.commands_loader:
            self._log.append(f"  Cmd override: {part}")
            best_cmd, matched_tokens, operands = \
                self.commands_loader.find_command_for_input(part)
            if best_cmd and best_cmd.script:
                self._log.append(f"Command: {' '.join(matched_tokens)}")
                executor = ScriptExecutor(self.client)
                result = executor.execute(best_cmd.script, operands)
                return {
                    "command": ' '.join(matched_tokens),
                    "commands_sent": result.get("commands_sent", []),
                    "datarefs_set": result.get("datarefs_set", []),
                    "errors": result.get("errors", []),
                }
        return None

    def _try_execute_command(self, item: ChecklistItem) -> Optional[Dict]:
        if not self.commands_loader:
            return None
        from .script_executor import ScriptExecutor

        # Use command_override if set (from CSV/XLSX)
        if item.command_override:
            # Support multiple commands separated by ;
            parts = [p for p in item.command_override.split(";") if p.strip()]
            if len(parts) == 1:
                return self._execute_single_command(parts[0])

            merged = {
                "command": item.command_override,
                "commands_sent": [],
                "datarefs_set": [],
                "errors": [],
            }
            any_result = False
            for part in parts:
                result = self._execute_single_command(part)
                if result:
                    any_result = True
                    merged["commands_sent"].extend(result.get("commands_sent", []))
                    merged["datarefs_set"].extend(result.get("datarefs_set", []))
                    merged["errors"].extend(result.get("errors", []))
            return merged if any_result else None

        # Exact lookup: "{label} {status_text}" in commands.xml — no guessing
        text = f"{item.label} {item.checked_text}".strip()
        self._log.append(f"  Cmd lookup: {text!r}")
        best_cmd, matched_tokens, operands = \
            self.commands_loader.find_command_for_input(text)
        if best_cmd and best_cmd.script:
            self._log.append(f"Command: {' '.join(matched_tokens)}")
            executor = ScriptExecutor(self.client)
            result = executor.execute(best_cmd.script, operands)
            return {
                "command": ' '.join(matched_tokens),
                "commands_sent": result.get("commands_sent", []),
                "datarefs_set": result.get("datarefs_set", []),
                "errors": result.get("errors", []),
            }
        return None

    # Zibo-specific post-command actions for items where commands.xml is incomplete.
    # The BATTERY ON command only opens the cover — it doesn't flip the switch.
    _ZIBO_POST_COMMAND = {
        "BATTERY": {
            "ON": [("laminar/B738/electric/battery_pos", 1)],
            "OFF": [("laminar/B738/electric/battery_pos", 0)],
        },
        "YAW DAMPER": {
            "ON": [("laminar/B738/toggle_switch/yaw_dumper_pos", 1)],
            "OFF": [("laminar/B738/toggle_switch/yaw_dumper_pos", 0)],
        },
        "IRS MODE SELECTORS": {
            "NAV": "irs_nav",
        },
        # Fuel pumps — toggle switches, direct writes rejected
        "FUEL PUMPS LEFT AND RIGHT": {
            "ON": [
                ("cmd:laminar/B738/toggle_switch/fuel_pump_lft1", None),
                ("cmd:laminar/B738/toggle_switch/fuel_pump_lft2", None),
                ("cmd:laminar/B738/toggle_switch/fuel_pump_rgt1", None),
                ("cmd:laminar/B738/toggle_switch/fuel_pump_rgt2", None),
            ],
        },
        "FUEL PUMPS CENTER": {
            "ON": [
                ("cmd:laminar/B738/toggle_switch/fuel_pump_ctr1", None),
                ("cmd:laminar/B738/toggle_switch/fuel_pump_ctr2", None),
            ],
        },
        # Window heaters — toggle switches
        "WINDOW HEATERS": {
            "ON": [
                ("cmd:laminar/B738/toggle_switch/window_heat_l_side", None),
                ("cmd:laminar/B738/toggle_switch/window_heat_l_fwd", None),
                ("cmd:laminar/B738/toggle_switch/window_heat_r_fwd", None),
                ("cmd:laminar/B738/toggle_switch/window_heat_r_side", None),
            ],
        },
        "WINDOW HEATER": {
            "ON": [
                ("cmd:laminar/B738/toggle_switch/window_heat_l_side", None),
                ("cmd:laminar/B738/toggle_switch/window_heat_l_fwd", None),
                ("cmd:laminar/B738/toggle_switch/window_heat_r_fwd", None),
                ("cmd:laminar/B738/toggle_switch/window_heat_r_side", None),
            ],
        },
        # Emergency exit lights — toggle command sequence
        "EMERGENCY EXIT LIGHTS": {
            "ARMED": [
                ("cmd:laminar/B738/button_switch_cover09", None),       # open cover
                ("cmd:laminar/B738/push_button/emer_exit_full_off", None),  # reset to known state
                ("cmd:laminar/B738/button_switch_cover09", None),       # close cover = ARMED
            ],
        },
        # Hydraulic pump — direct dataref write works
        # (removed toggle — _write_conditions handles it)
        # Engine start switches — rotary commands
        "ENGINE 1 START SWITCH": {
            "GROUND": [("cmd:laminar/B738/rotary/eng1_start_grd", None)],
        },
        "ENGINE 2 START SWITCH": {
            "GROUND": [("cmd:laminar/B738/rotary/eng2_start_grd", None)],
        },
        # Mixture levers — direct dataref write to idle (1.0)
        "ENGINE 1 START LEVER": {
            "IDLE": [("laminar/B738/engine/mixture_ratio1", 1)],
        },
        "ENGINE 2 START LEVER": {
            "IDLE": [("laminar/B738/engine/mixture_ratio2", 1)],
        },
    }

    def _zibo_post_command(self, item: ChecklistItem):
        """Apply Zibo-specific dataref writes for incomplete commands."""
        label = item.label.strip().upper()
        checked = item.checked_text.strip().upper()
        mapping = self._ZIBO_POST_COMMAND.get(label)
        if not mapping:
            return
        checked_key = checked.split(',')[0].split('AND')[0].strip()
        writes = mapping.get(checked_key)
        if not writes:
            return
        if isinstance(writes, str):
            if writes == "irs_nav":
                self._step_irs_to_nav()
            return
        if isinstance(writes, list):
            # Check if conditions already met — skip toggles to avoid flipping back
            if self._check_all_conditions(item.conditions, item.condition_logic):
                self._log.append(f"  Zibo: already satisfied, skip toggles")
                return
            for entry, value in writes:
                try:
                    if isinstance(entry, str) and entry.startswith("cmd:"):
                        cmd_name = entry[4:]
                        self.client.send_command(cmd_name)
                        self._log.append(f"  Zibo cmd: {cmd_name}")
                    else:
                        self.client.set_dataref(entry, value)
                        self._log.append(f"  Zibo fix: {entry} = {value}")
                    time.sleep(1.0)
                except Exception as e:
                    self._log.append(f"  Zibo fix failed: {entry} - {e}")

    def _toggle_if_needed(self, status_dataref: str, target: float, toggle_cmd: str):
        """Send a toggle command only if the current state doesn't match target."""
        current = self._read_dataref(status_dataref)
        try:
            val = float(current) if current is not None else -1
        except (TypeError, ValueError):
            val = -1
        if abs(val - target) < 0.05:
            self._log.append(f"  Already at {target}: {status_dataref}")
            return
        try:
            self.client.send_command(toggle_cmd)
            self._log.append(f"  Zibo toggle: {toggle_cmd}")
            time.sleep(0.3)
        except Exception as e:
            self._log.append(f"  Zibo toggle failed: {toggle_cmd} - {e}")

    def _step_irs_to_nav(self):
        """Step both IRS rotary knobs to NAV (position 2) using commands."""
        knobs = [
            ("laminar/B738/toggle_switch/irs_left", "laminar/B738/toggle_switch/irs_L_right"),
            ("laminar/B738/toggle_switch/irs_right", "laminar/B738/toggle_switch/irs_R_right"),
        ]
        for dataref, cmd_right in knobs:
            current = self._read_dataref(dataref)
            try:
                pos = int(float(current)) if current is not None else 0
            except (TypeError, ValueError):
                pos = 0
            steps = 2 - pos  # NAV = 2
            if steps <= 0:
                self._log.append(f"  IRS {dataref} already at {pos}")
                continue
            for _ in range(steps):
                try:
                    self.client.send_command(cmd_right)
                    self._log.append(f"  IRS step: {cmd_right}")
                    time.sleep(0.3)
                except Exception as e:
                    self._log.append(f"  IRS step failed: {cmd_right} - {e}")

    def _try_toggle_unsatisfied(self, conditions: List[DatarefCondition]):
        """For unsatisfied conditions involving cover/switch positions, try toggling."""
        for cond in conditions:
            if cond.operator.startswith('+'):
                continue
            raw = self._read_dataref(cond.dataref, cond.index)
            if _check_condition(cond, raw):
                continue
            # cover_position datarefs need toggle commands, not direct writes
            if 'cover_position' in cond.dataref and cond.index is not None:
                # Zibo 737 cover commands follow pattern: laminar/B738/button_switch_coverNN
                # cover index mapping: index 2 -> cover02, index 3 -> cover03, etc.
                cover_cmd = f"laminar/B738/button_switch_cover{cond.index:02d}"
                self._log.append(f"Toggling cover: {cover_cmd}")
                try:
                    self.client.send_command(cover_cmd)
                    time.sleep(0.3)
                except Exception as e:
                    self._log.append(f"Cover toggle failed: {e}")

    # Datarefs that represent simulator-computed STATUS, not pilot-controllable switches.
    # These must NEVER be written — only the sim sets them based on physical state.
    # Writing these fakes condition satisfaction (e.g. APU bus shows hot when it's not).
    _READ_ONLY_DATAREFS = {
        # APU status — sim computes from APU physical state
        "sim/cockpit/engine/APU_running",
        "laminar/B738/electrical/apu_bus_enable",
        # Annunciators — sim-driven indicator lights
        "laminar/B738/annunciator/source_off1",
        "laminar/B738/annunciator/source_off2",
        "laminar/B738/annunciator/hyd_stdby_rud",
        # Failure states
        "sim/operation/failures/rel_batter0",
        # Switches actuated by toggle commands — direct writes rejected by Zibo
        # yaw_damper_on — _write_conditions will try; _ZIBO_POST_COMMAND writes yaw_dumper_pos
        "laminar/B738/toggle_switch/emer_exit_lights",
        "laminar/B738/ice/window_heat_l_side_pos",
        "laminar/B738/ice/window_heat_l_fwd_pos",
        "laminar/B738/ice/window_heat_r_fwd_pos",
        "laminar/B738/ice/window_heat_r_side_pos",
        "laminar/B738/fuel/fuel_tank_pos_lft1",
        "laminar/B738/fuel/fuel_tank_pos_lft2",
        "laminar/B738/fuel/fuel_tank_pos_rgt1",
        "laminar/B738/fuel/fuel_tank_pos_rgt2",
        "laminar/B738/fuel/fuel_tank_pos_ctr1",
        "laminar/B738/fuel/fuel_tank_pos_ctr2",
        "laminar/B738/engine/starter1_pos",
        "laminar/B738/engine/starter2_pos",
        # mixture_ratio1/2 — writable, set directly to 1.0 for idle
        # IRS alignment status — sim-computed, takes minutes to align
        "laminar/B738/irs/irs_mode",
        "laminar/B738/irs/irs2_mode",
        # Engine parameters — sim-computed
        "sim/cockpit2/engine/indicators/N1_percent",
        "sim/cockpit2/engine/indicators/N2_percent",
        "sim/cockpit2/engine/indicators/oil_pressure_psi",
        "laminar/B738/engine/indicators/N1_percent_1",
        "laminar/B738/engine/indicators/N1_percent_2",
        "laminar/B738/engine/indicators/N2_percent_1",
        "laminar/B738/engine/indicators/N2_percent_2",
    }

    def _is_read_only(self, dataref: str) -> bool:
        """Check if a dataref is sim-computed status that must not be written."""
        if dataref in self._READ_ONLY_DATAREFS:
            return True
        # Block entire annunciator and failure subtrees
        if "annunciator" in dataref:
            return True
        if "failures" in dataref:
            return True
        # Block engine indicator subtrees
        if "indicators/" in dataref:
            return True
        return False

    def _write_conditions(self, conditions: List[DatarefCondition]) -> Tuple[List[Dict], List[str]]:
        """Write condition datarefs to actuate switches — but NEVER status datarefs."""
        actions = []
        errors = []
        for cond in conditions:
            if cond.operator.startswith('+'):
                continue
            if self._is_read_only(cond.dataref):
                continue
            actual = self._resolve_cross_ref(cond)
            target = _compute_write_value(actual)
            if target is None:
                continue
            try:
                if cond.index is not None:
                    ref = f"{cond.dataref}[{cond.index}]"
                    self.client.set_dataref(ref, target)
                    actions.append({"dataref": ref, "value": target})
                else:
                    self.client.set_dataref(cond.dataref, target)
                    actions.append({"dataref": cond.dataref, "value": target})
                self._log.append(f"  Write: {cond.dataref} = {target}")
            except Exception as e:
                errors.append(f"Failed to set {cond.dataref}: {e}")
        return actions, errors

    def _read_dataref(self, dataref: str, index: Optional[int] = None) -> Any:
        val = self.client.get_dataref(dataref, timeout=1.0)
        if val is not None and index is not None and isinstance(val, list):
            if index < len(val):
                return val[index]
            return None
        return val
